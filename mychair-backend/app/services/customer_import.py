"""
Bulk customer (client) import from CSV / Excel files.

Supports streaming-friendly batch inserts, phone normalisation, duplicate
skipping, and structured error reporting — without bypassing tenant isolation
or Beanie insert hooks (created_by / updated_by / timestamps).
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from email_validator import EmailNotValidError, validate_email
from openpyxl import Workbook, load_workbook

from app.models.audit import AuditLog
from app.models.customer import Customer
from app.models.user import User
from app.services.notifications import notification_service
from app.utils.timezone import now_utc

logger = logging.getLogger("customer_import")

BATCH_SIZE = 500
MAX_FILE_BYTES = 20 * 1024 * 1024  # 20 MB
MAX_ROWS = 50_000

TEMPLATE_HEADERS: Tuple[str, ...] = (
    "Full Name",
    "Mobile Number",
    "Email",
    "Gender",
    "DOB",
    "Anniversary",
    "Address",
    "City",
    "State",
    "Country",
    "Pincode",
    "Notes",
)

# Normalised header → canonical key
_HEADER_ALIASES: Dict[str, str] = {
    "full name": "full_name",
    "fullname": "full_name",
    "client name": "full_name",
    "customer name": "full_name",
    "mobile number": "mobile",
    "mobile": "mobile",
    "phone": "mobile",
    "phone number": "mobile",
    "email": "email",
    "gender": "gender",
    "dob": "dob",
    "date of birth": "dob",
    "anniversary": "anniversary",
    "address": "address",
    "city": "city",
    "state": "state",
    "country": "country",
    "pincode": "pincode",
    "pin code": "pincode",
    "zip": "pincode",
    "zipcode": "pincode",
    "notes": "notes",
    "note": "notes",
}

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

REASON_DUPLICATE = "Duplicate Mobile"
REASON_MISSING_NAME = "Missing Name"
REASON_MISSING_MOBILE = "Missing Mobile"
REASON_INVALID_MOBILE = "Invalid Mobile"
REASON_INVALID_DATA = "Invalid Data"
REASON_INVALID_EMAIL = "Invalid Email"
REASON_INVALID_DOB = "Invalid DOB"
REASON_INVALID_ANNIVERSARY = "Invalid Anniversary"
REASON_INVALID_GENDER = "Invalid Gender"


@dataclass
class RowError:
    row: int
    mobile: Optional[str]
    reason: str
    status: str  # skipped | failed
    full_name: Optional[str] = None
    original: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "row": self.row,
            "mobile": self.mobile,
            "reason": self.reason,
            "status": self.status,
            "full_name": self.full_name,
            "original": self.original,
        }


@dataclass
class ImportResult:
    total_rows: int = 0
    inserted: int = 0
    duplicates: int = 0
    failed: int = 0
    errors: List[RowError] = field(default_factory=list)
    reasons: Dict[str, int] = field(default_factory=dict)

    def bump_reason(self, reason: str) -> None:
        self.reasons[reason] = self.reasons.get(reason, 0) + 1

    def to_dict(self) -> Dict[str, Any]:
        skipped = self.duplicates
        return {
            "totalRows": self.total_rows,
            "inserted": self.inserted,
            "duplicates": self.duplicates,
            "skipped": skipped,
            "failed": self.failed,
            "errors": [e.to_dict() for e in self.errors],
            "reasons": self.reasons,
        }


class CustomerImportError(Exception):
    """Raised for file-level (not row-level) import failures."""

    def __init__(self, message: str, status_code: int = 400):
        super().__init__(message)
        self.message = message
        self.status_code = status_code


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def normalize_mobile(raw: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Strip spaces/dashes/brackets and optional +91 / 91 / leading 0.
    Returns (normalized_digits, error_reason).
    """
    if raw is None:
        return None, REASON_MISSING_MOBILE
    text = str(raw).strip()
    if not text:
        return None, REASON_MISSING_MOBILE

    # Excel may give floats like 9876543210.0
    if isinstance(raw, float) and raw == int(raw):
        text = str(int(raw))
    elif re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]

    cleaned = re.sub(r"[\s\-\(\)\.]", "", text)
    if cleaned.startswith("+"):
        cleaned = cleaned[1:]
    if cleaned.startswith("91") and len(cleaned) > 10:
        cleaned = cleaned[2:]
    if cleaned.startswith("0") and len(cleaned) == 11:
        cleaned = cleaned[1:]

    if not cleaned.isdigit():
        return None, REASON_INVALID_MOBILE
    if len(cleaned) < 10 or len(cleaned) > 15:
        return None, REASON_INVALID_MOBILE
    return cleaned, None


def split_full_name(raw: Any) -> Tuple[Optional[str], str, Optional[str]]:
    """Returns (first_name, last_name, error_reason)."""
    if raw is None:
        return None, "", REASON_MISSING_NAME
    text = str(raw).strip()
    if not text:
        return None, "", REASON_MISSING_NAME
    parts = text.split(None, 1)
    first = parts[0][:50]
    last = (parts[1][:50] if len(parts) > 1 else "")
    return first, last, None


def parse_import_date(raw: Any) -> Tuple[Optional[datetime], Optional[str]]:
    """
    Accept DD/MM/YYYY (preferred for salon owners), YYYY-MM-DD, DD-MM-YYYY,
    or Excel datetime cells. Returns (datetime|None, error_message|None).
    Empty → (None, None).
    """
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, None
    if isinstance(raw, datetime):
        return raw.replace(tzinfo=None), None
    # openpyxl may return date-like without time
    if hasattr(raw, "year") and hasattr(raw, "month") and hasattr(raw, "day") and not isinstance(raw, str):
        try:
            return datetime(int(raw.year), int(raw.month), int(raw.day)), None
        except Exception:
            return None, "Invalid date"

    text = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt), None
        except ValueError:
            continue
    return None, "Invalid date"


def normalize_gender(raw: Any) -> Tuple[Optional[str], Optional[str]]:
    if raw is None or (isinstance(raw, str) and not str(raw).strip()):
        return None, None
    text = str(raw).strip().upper()
    mapping = {
        "M": "MALE",
        "MALE": "MALE",
        "MAN": "MALE",
        "F": "FEMALE",
        "FEMALE": "FEMALE",
        "WOMAN": "FEMALE",
        "O": "OTHER",
        "OTHER": "OTHER",
        "OTHERS": "OTHER",
        "NON-BINARY": "OTHER",
        "NONBINARY": "OTHER",
    }
    value = mapping.get(text)
    if not value:
        return None, REASON_INVALID_GENDER
    return value, None


def sanitize_optional_str(raw: Any, max_len: int = 500) -> Optional[str]:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    # Strip control characters
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    return text[:max_len]


def normalize_email(raw: Any) -> Tuple[Optional[str], Optional[str]]:
    text = sanitize_optional_str(raw, 254)
    if not text:
        return None, None
    try:
        result = validate_email(text, check_deliverability=False)
        return result.normalized, None
    except EmailNotValidError:
        return None, REASON_INVALID_EMAIL


def map_headers(raw_headers: Sequence[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, header in enumerate(raw_headers):
        key = _HEADER_ALIASES.get(_norm_header(header))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _cell(row: Sequence[Any], mapping: Dict[str, int], key: str) -> Any:
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _row_original(row: Sequence[Any], mapping: Dict[str, int]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for key in (
        "full_name",
        "mobile",
        "email",
        "gender",
        "dob",
        "anniversary",
        "address",
        "city",
        "state",
        "country",
        "pincode",
        "notes",
    ):
        val = _cell(row, mapping, key)
        if val is None or (isinstance(val, str) and not val.strip()):
            out[key] = ""
        elif isinstance(val, datetime):
            out[key] = val.strftime("%d/%m/%Y")
        else:
            out[key] = str(val).strip()
    return out


def detect_extension(filename: str) -> str:
    name = (filename or "").lower().strip()
    for ext in (".xlsx", ".xls", ".csv"):
        if name.endswith(ext):
            return ext
    return ""


def validate_upload_file(
    filename: str,
    content_type: Optional[str],
    size: int,
) -> str:
    if size <= 0:
        raise CustomerImportError("Uploaded file is empty.")
    if size > MAX_FILE_BYTES:
        raise CustomerImportError("File exceeds the maximum size of 20 MB.")

    ext = detect_extension(filename)
    if ext not in ALLOWED_EXTENSIONS:
        raise CustomerImportError(
            "Unsupported file type. Please upload a .csv, .xlsx, or .xls file."
        )

    if content_type:
        mime = content_type.split(";")[0].strip().lower()
        if any(x in mime for x in ("executable", "javascript", "x-msdownload", "x-sh", "x-msdos")):
            raise CustomerImportError("Unsupported or unsafe file type.")

    return ext


def sniff_magic(content: bytes, ext: str) -> None:
    """Reject renamed executables / obviously wrong payloads."""
    if not content:
        raise CustomerImportError("Uploaded file is empty.")
    # PE / ELF executables
    if content[:2] == b"MZ" or content[:4] == b"\x7fELF":
        raise CustomerImportError("Unsupported or unsafe file type.")
    if ext == ".xlsx":
        # ZIP/OOXML
        if content[:2] != b"PK":
            raise CustomerImportError("File content does not match a valid .xlsx workbook.")
    elif ext == ".xls":
        # OLE compound document
        if content[:4] != b"\xd0\xcf\x11\xe0":
            raise CustomerImportError("File content does not match a valid .xls workbook.")


def parse_rows_from_bytes(content: bytes, ext: str) -> Tuple[Dict[str, int], List[List[Any]]]:
    if ext == ".csv":
        return _parse_csv(content)
    if ext == ".xlsx":
        return _parse_xlsx(content)
    if ext == ".xls":
        return _parse_xls(content)
    raise CustomerImportError("Unsupported file type.")


def _parse_csv(content: bytes) -> Tuple[Dict[str, int], List[List[Any]]]:
    # Try utf-8-sig (BOM), then latin-1 fallback
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise CustomerImportError("Could not decode CSV file. Please use UTF-8 encoding.")

    reader = csv.reader(io.StringIO(text))
    try:
        header_row = next(reader)
    except StopIteration:
        raise CustomerImportError("File has no header row.") from None

    mapping = map_headers(header_row)
    if "full_name" not in mapping or "mobile" not in mapping:
        raise CustomerImportError(
            "Invalid headers. Template must include 'Full Name' and 'Mobile Number' columns."
        )

    rows: List[List[Any]] = []
    for row in reader:
        if not row or all(not str(c).strip() for c in row):
            continue
        rows.append(row)
        if len(rows) > MAX_ROWS:
            raise CustomerImportError(f"File exceeds the maximum of {MAX_ROWS:,} data rows.")
    return mapping, rows


def _parse_xlsx(content: bytes) -> Tuple[Dict[str, int], List[List[Any]]]:
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise CustomerImportError("Could not read Excel (.xlsx) file.") from exc

    # Prefer sheet named Clients / Data / first sheet (skip Example)
    sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() in ("clients", "data", "template", "customers"):
            sheet = wb[name]
            break
    if sheet is None:
        for name in wb.sheetnames:
            if name.strip().lower() != "example":
                sheet = wb[name]
                break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise CustomerImportError("File has no header row.") from None

    mapping = map_headers(list(header_row or []))
    if "full_name" not in mapping or "mobile" not in mapping:
        raise CustomerImportError(
            "Invalid headers. Template must include 'Full Name' and 'Mobile Number' columns."
        )

    rows: List[List[Any]] = []
    for row in rows_iter:
        values = list(row or [])
        if not values or all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        rows.append(values)
        if len(rows) > MAX_ROWS:
            raise CustomerImportError(f"File exceeds the maximum of {MAX_ROWS:,} data rows.")
    wb.close()
    return mapping, rows


def _parse_xls(content: bytes) -> Tuple[Dict[str, int], List[List[Any]]]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise CustomerImportError(
            "XLS support is not available on this server. Please upload .xlsx or .csv."
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise CustomerImportError("Could not read Excel (.xls) file.") from exc

    sheet = book.sheet_by_index(0)
    if sheet.nrows < 1:
        raise CustomerImportError("File has no header row.")

    header_row = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
    mapping = map_headers(header_row)
    if "full_name" not in mapping or "mobile" not in mapping:
        raise CustomerImportError(
            "Invalid headers. Template must include 'Full Name' and 'Mobile Number' columns."
        )

    rows: List[List[Any]] = []
    for r in range(1, sheet.nrows):
        values = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            val = cell.value
            # xlrd date cells
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    val = datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
                except Exception:
                    pass
            values.append(val)
        if all(v is None or (isinstance(v, str) and not str(v).strip()) or v == "" for v in values):
            continue
        rows.append(values)
        if len(rows) > MAX_ROWS:
            raise CustomerImportError(f"File exceeds the maximum of {MAX_ROWS:,} data rows.")
    return mapping, rows


@dataclass
class _ValidatedRow:
    row_number: int  # 1-based data row (header = row 1 → first data = 2)
    first_name: str
    last_name: str
    phone: str
    email: Optional[str]
    gender: Optional[str]
    dob: Optional[datetime]
    address: Optional[str]
    notes: Optional[str]
    metadata: Dict[str, Any]
    original: Dict[str, Any]


def validate_data_rows(
    mapping: Dict[str, int],
    rows: List[List[Any]],
) -> Tuple[List[_ValidatedRow], List[RowError], Dict[str, int]]:
    valid: List[_ValidatedRow] = []
    errors: List[RowError] = []
    reasons: Dict[str, int] = {}
    seen_phones: Set[str] = set()

    def bump(reason: str) -> None:
        reasons[reason] = reasons.get(reason, 0) + 1

    for offset, row in enumerate(rows):
        row_number = offset + 2  # header is row 1
        original = _row_original(row, mapping)

        first, last, name_err = split_full_name(_cell(row, mapping, "full_name"))
        phone, phone_err = normalize_mobile(_cell(row, mapping, "mobile"))

        if name_err:
            errors.append(
                RowError(
                    row=row_number,
                    mobile=phone or original.get("mobile") or None,
                    reason=name_err,
                    status="failed",
                    full_name=original.get("full_name") or None,
                    original=original,
                )
            )
            bump(name_err)
            continue

        if phone_err:
            errors.append(
                RowError(
                    row=row_number,
                    mobile=original.get("mobile") or None,
                    reason=phone_err,
                    status="failed",
                    full_name=f"{first} {last}".strip(),
                    original=original,
                )
            )
            bump(phone_err)
            continue

        assert first is not None and phone is not None

        if phone in seen_phones:
            errors.append(
                RowError(
                    row=row_number,
                    mobile=phone,
                    reason=REASON_DUPLICATE,
                    status="skipped",
                    full_name=f"{first} {last}".strip(),
                    original=original,
                )
            )
            bump(REASON_DUPLICATE)
            continue

        email, email_err = normalize_email(_cell(row, mapping, "email"))
        if email_err:
            errors.append(
                RowError(
                    row=row_number,
                    mobile=phone,
                    reason=email_err,
                    status="failed",
                    full_name=f"{first} {last}".strip(),
                    original=original,
                )
            )
            bump(email_err)
            continue

        gender, gender_err = normalize_gender(_cell(row, mapping, "gender"))
        if gender_err:
            errors.append(
                RowError(
                    row=row_number,
                    mobile=phone,
                    reason=gender_err,
                    status="failed",
                    full_name=f"{first} {last}".strip(),
                    original=original,
                )
            )
            bump(gender_err)
            continue

        dob, dob_err = parse_import_date(_cell(row, mapping, "dob"))
        if dob_err:
            errors.append(
                RowError(
                    row=row_number,
                    mobile=phone,
                    reason=REASON_INVALID_DOB,
                    status="failed",
                    full_name=f"{first} {last}".strip(),
                    original=original,
                )
            )
            bump(REASON_INVALID_DOB)
            continue

        anniversary, ann_err = parse_import_date(_cell(row, mapping, "anniversary"))
        if ann_err:
            errors.append(
                RowError(
                    row=row_number,
                    mobile=phone,
                    reason=REASON_INVALID_ANNIVERSARY,
                    status="failed",
                    full_name=f"{first} {last}".strip(),
                    original=original,
                )
            )
            bump(REASON_INVALID_ANNIVERSARY)
            continue

        address = sanitize_optional_str(_cell(row, mapping, "address"), 500)
        city = sanitize_optional_str(_cell(row, mapping, "city"), 100)
        state = sanitize_optional_str(_cell(row, mapping, "state"), 100)
        country = sanitize_optional_str(_cell(row, mapping, "country"), 100)
        pincode = sanitize_optional_str(_cell(row, mapping, "pincode"), 20)
        notes = sanitize_optional_str(_cell(row, mapping, "notes"), 2000)

        # Compose display address if parts provided and address empty
        if not address:
            parts = [p for p in (city, state, country, pincode) if p]
            address = ", ".join(parts) if parts else None

        metadata: Dict[str, Any] = {}
        if anniversary:
            metadata["anniversary"] = anniversary.date().isoformat()
        if city:
            metadata["city"] = city
        if state:
            metadata["state"] = state
        if country:
            metadata["country"] = country
        if pincode:
            metadata["pincode"] = pincode

        seen_phones.add(phone)
        valid.append(
            _ValidatedRow(
                row_number=row_number,
                first_name=first,
                last_name=last,
                phone=phone,
                email=email,
                gender=gender,
                dob=dob,
                address=address,
                notes=notes,
                metadata=metadata,
                original=original,
            )
        )

    return valid, errors, reasons


async def _existing_phones(tenant_id: Optional[str], phones: Iterable[str]) -> Set[str]:
    phone_list = list(phones)
    existing: Set[str] = set()
    for i in range(0, len(phone_list), BATCH_SIZE):
        chunk = phone_list[i : i + BATCH_SIZE]
        query: Dict[str, Any] = {"phone": {"$in": chunk}, "is_deleted": False}
        if tenant_id:
            query["tenant_id"] = tenant_id
        docs = await Customer.find(query).to_list()
        for doc in docs:
            if doc.phone:
                existing.add(doc.phone)
    return existing


async def import_customers_from_file(
    *,
    content: bytes,
    filename: str,
    content_type: Optional[str],
    tenant_id: Optional[str],
    current_user: User,
) -> ImportResult:
    ext = validate_upload_file(filename, content_type, len(content))
    sniff_magic(content, ext)

    mapping, raw_rows = parse_rows_from_bytes(content, ext)
    if not raw_rows:
        raise CustomerImportError("File contains no data rows.")

    result = ImportResult(total_rows=len(raw_rows))
    valid_rows, row_errors, reasons = validate_data_rows(mapping, raw_rows)
    result.errors.extend(row_errors)
    result.reasons.update(reasons)
    result.failed = sum(1 for e in row_errors if e.status == "failed")
    result.duplicates = sum(1 for e in row_errors if e.status == "skipped")

    if not valid_rows:
        _log_import(current_user, tenant_id, result)
        return result

    existing = await _existing_phones(tenant_id, (r.phone for r in valid_rows))

    to_insert: List[_ValidatedRow] = []
    for vr in valid_rows:
        if vr.phone in existing:
            result.duplicates += 1
            result.bump_reason(REASON_DUPLICATE)
            result.errors.append(
                RowError(
                    row=vr.row_number,
                    mobile=vr.phone,
                    reason=REASON_DUPLICATE,
                    status="skipped",
                    full_name=f"{vr.first_name} {vr.last_name}".strip(),
                    original=vr.original,
                )
            )
            continue
        existing.add(vr.phone)  # prevent dupes within remaining insert set vs DB
        to_insert.append(vr)

    user_id = str(current_user.id) if current_user.id else None
    inserted = 0
    for i in range(0, len(to_insert), BATCH_SIZE):
        batch = to_insert[i : i + BATCH_SIZE]
        docs = [
            Customer(
                first_name=item.first_name,
                last_name=item.last_name,
                phone=item.phone,
                email=item.email,
                gender=item.gender,
                dob=item.dob,
                address=item.address,
                notes=item.notes,
                metadata=item.metadata or {},
                tenant_id=tenant_id,
                created_by=user_id,
                updated_by=user_id,
            )
            for item in batch
        ]
        try:
            await Customer.insert_many(docs)
            inserted += len(docs)
        except Exception:
            # Fallback: insert one-by-one so a single bad row doesn't lose the batch
            logger.exception(
                "Batch insert failed; falling back to per-row insert (tenant=%s)",
                tenant_id,
            )
            for item, doc in zip(batch, docs):
                try:
                    await doc.insert()
                    inserted += 1
                except Exception as row_exc:
                    logger.warning(
                        "Failed to insert customer phone=%s row=%s: %s",
                        item.phone,
                        item.row_number,
                        row_exc,
                    )
                    result.failed += 1
                    result.bump_reason(REASON_INVALID_DATA)
                    result.errors.append(
                        RowError(
                            row=item.row_number,
                            mobile=item.phone,
                            reason=REASON_INVALID_DATA,
                            status="failed",
                            full_name=f"{item.first_name} {item.last_name}".strip(),
                            original=item.original,
                        )
                    )

    result.inserted = inserted
    _log_import(current_user, tenant_id, result)

    # Single summary notification (avoid N notifications for bulk)
    if inserted > 0 and tenant_id:
        try:
            recipients = await notification_service._tenant_users_for_roles(
                tenant_id,
                tenant_id,
                ["salon_owner", "salon_admin", "salon_manager"],
            )
            await notification_service.create_event_notifications(
                tenant_id=tenant_id,
                salon_id=tenant_id,
                recipients=recipients,
                title="Bulk client import completed",
                body=(
                    f"{inserted} client{'s' if inserted != 1 else ''} imported"
                    f" ({result.duplicates} skipped, {result.failed} failed)."
                ),
                category="CUSTOMER",
                notification_type="CUSTOMER_BULK_IMPORT",
                source_event="CUSTOMER_BULK_IMPORT",
                metadata={
                    "inserted": inserted,
                    "duplicates": result.duplicates,
                    "failed": result.failed,
                    "total_rows": result.total_rows,
                },
            )
        except Exception:
            logger.exception("Failed to send bulk import notification")

        try:
            await AuditLog(
                tenant_id=tenant_id,
                user_id=user_id,
                action="IMPORT",
                entity_name="Customer",
                entity_id=None,
                after_state=result.to_dict(),
            ).insert()
        except Exception:
            logger.exception("Failed to write bulk import audit log")

    return result


def _log_import(current_user: User, tenant_id: Optional[str], result: ImportResult) -> None:
    logger.info(
        "customer_import uploader=%s salon=%s total=%s inserted=%s skipped=%s failed=%s",
        getattr(current_user, "id", None),
        tenant_id,
        result.total_rows,
        result.inserted,
        result.duplicates,
        result.failed,
    )


# ─────────────────────────── templates ──────────────────────────────────────

EXAMPLE_ROWS: Tuple[Tuple[str, ...], ...] = (
    (
        "Priya Sharma",
        "9876543210",
        "priya@example.com",
        "Female",
        "15/03/1990",
        "20/06/2015",
        "12 MG Road",
        "Bengaluru",
        "Karnataka",
        "India",
        "560001",
        "Prefers evening appointments",
    ),
    (
        "Rahul Verma",
        "9123456780",
        "",
        "Male",
        "01/01/1985",
        "",
        "",
        "Mumbai",
        "Maharashtra",
        "India",
        "400001",
        "",
    ),
)


def build_xlsx_template() -> bytes:
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.append(list(TEMPLATE_HEADERS))
    widths = [22, 16, 28, 10, 12, 14, 24, 14, 14, 12, 10, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    example = wb.create_sheet("Example")
    example.append(list(TEMPLATE_HEADERS))
    for row in EXAMPLE_ROWS:
        example.append(list(row))
    for idx, width in enumerate(widths, start=1):
        example.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv_template() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    # Include one blank guidance? Spec says first row headers only for main template.
    # Example rows are on the Example sheet for xlsx; for CSV append commented? Keep headers-only.
    return buf.getvalue().encode("utf-8-sig")


def build_error_report_csv(errors: Sequence[RowError]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "Row",
            "Status",
            "Error Message",
            "Full Name",
            "Mobile Number",
            "Email",
            "Gender",
            "DOB",
            "Anniversary",
            "Address",
            "City",
            "State",
            "Country",
            "Pincode",
            "Notes",
        ]
    )
    for err in errors:
        orig = err.original or {}
        writer.writerow(
            [
                err.row,
                err.status,
                err.reason,
                orig.get("full_name", err.full_name or ""),
                orig.get("mobile", err.mobile or ""),
                orig.get("email", ""),
                orig.get("gender", ""),
                orig.get("dob", ""),
                orig.get("anniversary", ""),
                orig.get("address", ""),
                orig.get("city", ""),
                orig.get("state", ""),
                orig.get("country", ""),
                orig.get("pincode", ""),
                orig.get("notes", ""),
            ]
        )
    return buf.getvalue().encode("utf-8-sig")
